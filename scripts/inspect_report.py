import openpyxl

def inspect():
    wb = openpyxl.load_workbook("E2E_Test_Report_VeriTask.xlsx")
    print("Sheets in workbook:", wb.sheetnames)
    
    # Check Dashboard
    ws_dash = wb["Summary Dashboard"]
    print("\n--- Dashboard Cells ---")
    for r in range(1, 14):
        row_vals = [ws_dash.cell(row=r, column=c).value for c in range(1, 7)]
        print(f"Row {r}: {row_vals}")
        
    # Check Details
    ws_details = wb["Test Cases Details"]
    print("\n--- Details Sheet Check ---")
    print("Total rows in details:", ws_details.max_row)
    print("Headers:", [ws_details.cell(row=1, column=c).value for c in range(1, 11)])
    print("First test case row:", [ws_details.cell(row=2, column=c).value for c in range(1, 11)])
    print("Last test case row:", [ws_details.cell(row=ws_details.max_row, column=c).value for c in range(1, 11)])

if __name__ == "__main__":
    inspect()
